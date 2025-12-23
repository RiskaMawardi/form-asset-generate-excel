import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
import os
import glob
import re
import requests
from io import BytesIO
from PIL import Image as PILImage
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

class SimpleExcelGenerator:
    def __init__(self, csv_file, template_file, output_folder='generated_excel', 
                 insert_images=True, send_email=False, email_config=None):
        self.csv_file = csv_file
        self.template_file = template_file
        self.output_folder = output_folder
        self.insert_images = insert_images
        self.send_email = send_email
        self.email_config = email_config or {}
        self.temp_image_folder = os.path.join(output_folder, 'temp_images')
        
        os.makedirs(output_folder, exist_ok=True)
        if insert_images:
            os.makedirs(self.temp_image_folder, exist_ok=True)

    def extract_year_from_asset_no(self, asset_no):
        """Extract year (19xx or 20xx) from asset number"""
        if not asset_no:
            return ''
        match = re.search(r'(19|20)\d{2}', str(asset_no))
        if match:
            return match.group(0)
        return ''
    
    def download_image_from_gdrive(self, url):
        """Download image from Google Drive URL"""
        try:
            file_id = None
            if 'id=' in url:
                file_id = url.split('id=')[1].split('&')[0]
            elif '/d/' in url:
                file_id = url.split('/d/')[1].split('/')[0]
            
            if not file_id:
                print(f"   WARNING: Could not extract file ID from URL: {url}")
                return None
            
            temp_filename = f"{file_id}.png"
            temp_path = os.path.join(self.temp_image_folder, temp_filename)
            
            if os.path.exists(temp_path):
                print(f"      -> Using cached image")
                return temp_path
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            }
            
            download_urls = [
                f"https://drive.google.com/uc?export=download&id={file_id}",
                f"https://drive.usercontent.google.com/download?id={file_id}&export=download",
                f"https://lh3.googleusercontent.com/d/{file_id}",
                f"https://drive.google.com/thumbnail?id={file_id}&sz=w1000"
            ]
            
            session = requests.Session()
            session.headers.update(headers)
            
            for attempt, download_url in enumerate(download_urls, 1):
                try:
                    print(f"      -> Attempt {attempt}/{len(download_urls)}: {download_url.split('?')[0]}...")
                    response = session.get(download_url, timeout=20, allow_redirects=True, verify=True)
                    
                    if response.status_code == 200 and len(response.content) > 1000:
                        content_type = response.headers.get('content-type', '').lower()
                        if 'text/html' in content_type:
                            print(f"      -> Got HTML response, trying next URL...")
                            continue
                        
                        try:
                            img = PILImage.open(BytesIO(response.content))
                            max_width = 200
                            if img.width > max_width:
                                ratio = max_width / img.width
                                new_height = int(img.height * ratio)
                                img = img.resize((max_width, new_height), PILImage.Resampling.LANCZOS)
                            
                            img.save(temp_path, 'PNG')
                            print(f"      -> ✓ Image downloaded successfully!")
                            return temp_path
                        except Exception as img_error:
                            print(f"      -> Invalid image data, trying next URL...")
                            continue
                    elif response.status_code == 404:
                        print(f"      -> File not found (404)")
                        break
                    else:
                        print(f"      -> HTTP {response.status_code}, trying next URL...")
                        continue
                except Exception as e:
                    print(f"      -> Error: {str(e)[:50]}...")
                    continue
            
            print(f"   ✗ All download attempts failed for this image")
            return None
        except Exception as e:
            print(f"   ✗ Unexpected error: {str(e)[:100]}")
            return None
    
    def read_csv_responses(self):
        """Read CSV file from Google Forms"""
        try:
            df = pd.read_csv(self.csv_file, encoding='utf-8')
            cols = pd.Series(df.columns)
            for dup in cols[cols.duplicated()].unique():
                indices = cols[cols == dup].index.values.tolist()
                for i, idx in enumerate(indices):
                    if i != 0:
                        cols[idx] = f"{dup}_duplicate{i}"
            df.columns = cols.str.strip()
            print(f"✓ Successfully read {len(df)} responses from CSV")
            return df
        except Exception as e:
            print(f"✗ Error reading CSV: {str(e)}")
            return None
    
    def extract_assets_from_row(self, row):
        """Extract all assets from a single row"""
        assets = []
        columns = row.index.tolist()
        
        asset_numbers = set()
        for col in columns:
            if 'upload' not in col.lower() and 'foto' not in col.lower():
                match = re.search(r'Asset\s+(\d+)', col, re.IGNORECASE)
                if match:
                    asset_numbers.add(int(match.group(1)))
        
        for num in sorted(asset_numbers):
            no_asset = None
            jenis = None
            foto = None
            
            for col in columns:
                if re.search(rf'\bAsset\s+{num}\b', col, re.IGNORECASE):
                    col_lower = col.lower()
                    
                    if 'no' in col_lower and 'upload' not in col_lower and 'foto' not in col_lower:
                        value = row[col]
                        if pd.notna(value) and str(value).strip() and not str(value).startswith('http'):
                            no_asset = str(value).strip()
                    elif 'jenis' in col_lower:
                        value = row[col]
                        if pd.notna(value) and str(value).strip():
                            jenis = str(value).strip()
                    elif 'upload' in col_lower or 'foto' in col_lower:
                        value = row[col]
                        if pd.notna(value) and str(value).strip():
                            foto = str(value).strip()
            
            if no_asset:
                assets.append({
                    'no': no_asset,
                    'jenis': jenis if jenis else '',
                    'foto': foto if foto else ''
                })
        
        if not assets:
            no_asset = None
            jenis = None
            foto = None
            
            for col in columns:
                col_lower = col.lower()
                if 'no' in col_lower and 'asset' in col_lower and 'upload' not in col_lower and 'foto' not in col_lower:
                    value = row[col]
                    if pd.notna(value) and str(value).strip() and not str(value).startswith('http'):
                        no_asset = str(value).strip()
                        break
            
            if no_asset:
                for col in columns:
                    col_lower = col.lower()
                    if 'jenis' in col_lower and 'asset' in col_lower:
                        value = row[col]
                        if pd.notna(value):
                            jenis = str(value).strip()
                            break
                
                for col in columns:
                    col_lower = col.lower()
                    if ('upload' in col_lower or 'foto' in col_lower) and 'asset' in col_lower:
                        value = row[col]
                        if pd.notna(value):
                            foto = str(value).strip()
                            break
                
                assets.append({
                    'no': no_asset,
                    'jenis': jenis if jenis else '',
                    'foto': foto if foto else ''
                })
        
        return assets
    
    def fill_excel_template(self, template_path, output_path, person_info, assets):
        """Fill Excel template with data"""
        try:
            wb = load_workbook(template_path)
            ws = wb.active

            timestamp = person_info.get('Timestamp', '')
            if timestamp:
                try:
                    if '/' in timestamp:
                        year = timestamp.split('/')[-1].split(' ')[0]
                    elif '-' in timestamp:
                        year = timestamp.split('-')[0]
                    else:
                        year = datetime.now().year
                    current_a1 = ws['E1'].value or "Tahun ..."
                    ws['E1'] = current_a1.replace('...', str(year))
                except:
                    current_a1 = ws['E1'].value or "Tahun ..."
                    ws['E1'] = current_a1.replace('...', str(datetime.now().year))
            
            ws['H1'] = person_info.get('Area', '')
            ws['H2'] = person_info.get('Divisi', '')
            ws['E27'] = person_info.get('Dibuat Oleh', '')
            ws['K28'] = person_info.get('PIC', '')
            
            for idx, asset in enumerate(assets):
                row_num = 9 + idx
                ws[f'A{row_num}'] = idx + 1
                ws[f'B{row_num}'] = asset['jenis']
                ws[f'C{row_num}'] = asset['no']
                ws[f'F{row_num}'] = person_info.get('Dipakai Oleh', '')
                ws[f'G{row_num}'] = person_info.get('Jabatan', '')
                ws[f'H{row_num}'] = person_info.get('Kondisi Inventaris', '')
                tahun_asset = self.extract_year_from_asset_no(asset['no'])
                ws[f'E{row_num}'] = tahun_asset
                
                if self.insert_images and asset['foto'] and asset['foto'].startswith('http'):
                    print(f"      -> Downloading image for asset {idx+1}...")
                    image_path = self.download_image_from_gdrive(asset['foto'])
                    
                    if image_path and os.path.exists(image_path):
                        try:
                            img = XLImage(image_path)
                            img.width = 100
                            img.height = 100
                            ws.add_image(img, f'D{row_num}')
                            ws.row_dimensions[row_num].height = 75
                            print(f"      -> Image inserted successfully")
                        except Exception as e:
                            print(f"      -> WARNING: Could not insert image: {str(e)}")
            
            wb.save(output_path)
            return True
        except Exception as e:
            print(f"ERROR filling template: {str(e)}")
            return False
    
    def send_email_with_attachments(self, recipient_email, person_name, files_data):
        """
        Send email with multiple Excel attachments
        files_data: list of dict [{'excel': path, 'asset_count': n}, ...]
        """
        try:
            # Validate email config
            if not all(k in self.email_config for k in ['smtp_server', 'smtp_port', 'sender_email', 'sender_password']):
                print(f"      -> ERROR: Email config incomplete")
                return False
            
            # Create message
            msg = MIMEMultipart()
            msg['From'] = self.email_config['sender_email']
            msg['To'] = recipient_email
            msg['Subject'] = f"Daftar Inventaris IT Asset - {person_name}"
            
            # Count total files
            total_excel = len(files_data)
            total_assets = sum(f.get('asset_count', 0) for f in files_data)
            
            # Email body
            body = f"""Halo {person_name},

Terlampir adalah daftar inventaris IT Asset Anda.

Total file terlampir:
- {total_excel} file Excel
- Total {total_assets} asset

Detail file:
"""
            
            for idx, file_info in enumerate(files_data, 1):
                asset_count = file_info.get('asset_count', 0)
                body += f"  {idx}. {asset_count} asset(s)\n"
            
            body += """
Jika ada pertanyaan, silakan hubungi tim IT.

Terima kasih,
IT Asset Management Team
"""
            msg.attach(MIMEText(body, 'plain'))
            
            # Attach all Excel files
            for idx, file_info in enumerate(files_data, 1):
                excel_path = file_info.get('excel')
                if excel_path and os.path.exists(excel_path):
                    with open(excel_path, 'rb') as f:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(f.read())
                        encoders.encode_base64(part)
                        # Add sequence number if multiple files
                        filename = os.path.basename(excel_path)
                        if len(files_data) > 1:
                            filename = f"{idx}_{filename}"
                        part.add_header('Content-Disposition', f'attachment; filename={filename}')
                        msg.attach(part)
            
            # Send email
            server = smtplib.SMTP(self.email_config['smtp_server'], self.email_config['smtp_port'])
            server.starttls()
            server.login(self.email_config['sender_email'], self.email_config['sender_password'])
            server.send_message(msg)
            server.quit()
            
            print(f"      -> ✓ Email sent to {recipient_email} ({total_excel} Excel)")
            return True
            
        except Exception as e:
            print(f"      -> ✗ Email failed: {str(e)}")
            return False
    
    def generate_excel_consolidated(self):
        """Generate one Excel per person with all their assets"""
        df = self.read_csv_responses()
        if df is None or len(df) == 0:
            print("No data to process!")
            return
        
        # Normalize column names
        column_mapping = {}
        for col in df.columns:
            col_lower = col.lower().strip()
            
            if 'timestamp' in col_lower:
                column_mapping[col] = 'Timestamp'
            elif 'email' in col_lower and 'address' in col_lower:
                column_mapping[col] = 'Email'
            elif col_lower == 'dipakai oleh':
                column_mapping[col] = 'Dipakai Oleh'
            elif col_lower == 'dibuat oleh':
                column_mapping[col] = 'Dibuat Oleh'
            elif col_lower == 'jabatan':
                column_mapping[col] = 'Jabatan'
            elif col_lower == 'divisi':
                column_mapping[col] = 'Divisi'
            elif col_lower == 'area':
                column_mapping[col] = 'Area'
            elif col_lower == 'pic':
                column_mapping[col] = 'PIC'
        
        df = df.rename(columns=column_mapping)
        
        # Group by person
        df['person_key'] = (
            df.get('Dipakai Oleh', 'Unknown').fillna('Unknown') + '_' +
            df.get('Divisi', 'Unknown').fillna('Unknown') + '_' +
            df.get('Area', 'Unknown').fillna('Unknown')
        )
        
        print(f"\nProcessing {len(df)} responses...\n")
        
        grouped = {}
        for _, row in df.iterrows():
            person_key = row['person_key']
            
            if person_key not in grouped:
                # Get Kondisi Inventaris - try different columns
                kondisi_value = ''
                # Try base column first (for single asset)
                if 'Kondisi Inventaris' in row.index:
                    kondisi_value = row.get('Kondisi Inventaris', '')
                # If empty, try numbered columns
                if not kondisi_value or pd.isna(kondisi_value):
                    for i in range(1, 10):
                        col_name = f'Kondisi Inventaris {i}'
                        if col_name in row.index:
                            val = row.get(col_name, '')
                            if val and not pd.isna(val):
                                kondisi_value = val
                                break
                
                # Convert to string safely
                if pd.isna(kondisi_value):
                    kondisi_value = ''
                else:
                    kondisi_value = str(kondisi_value).strip()
                
                grouped[person_key] = {
                    'info': {
                        'Timestamp': row.get('Timestamp', ''),
                        'Dipakai Oleh': row.get('Dipakai Oleh', 'Unknown'),
                        'Divisi': row.get('Divisi', 'Unknown'),
                        'Area': row.get('Area', 'Unknown'),
                        'PIC': row.get('PIC', 'Unknown'),
                        'Dibuat Oleh': row.get('Dibuat Oleh', ''),
                        'Jabatan': row.get('Jabatan', ''),
                        'Kondisi Inventaris': kondisi_value,
                        'Email': row.get('Email', '')
                    },
                    'assets': []
                }
            
            assets = self.extract_assets_from_row(row)
            grouped[person_key]['assets'].extend(assets)
        
        # Generate files and send emails
        success = 0
        email_sent = 0
        
        # Group files by email for sending
        email_groups = {}  # {email: [{'excel': path, 'pdf': path, 'asset_count': n, 'name': str}, ...]}
        
        for idx, (person_key, data) in enumerate(grouped.items(), 1):
            info = data['info']
            assets = data['assets']
            
            nama = str(info['Dipakai Oleh']).replace(' ', '_')
            area = str(info['Area']).replace(' ', '_').replace('(', '').replace(')', '').replace('.', '')
            divisi = str(info['Divisi']).replace(' ', '_')
            
            filename = f"{idx}_{area}_{divisi}_{nama}_{len(assets)}assets.xlsx"
            output_path = os.path.join(self.output_folder, filename)
            
            if self.fill_excel_template(self.template_file, output_path, info, assets):
                print(f"OK [{idx}/{len(grouped)}] Created: {filename}")
                print(f"   -> {len(assets)} asset(s)")
                for i, asset in enumerate(assets, 1):
                    print(f"      {i}. {asset['jenis']} | {asset['no']}")
                
                # Collect files for email grouping
                if self.send_email and info.get('Email'):
                    recipient_email = info['Email'].strip()
                    if recipient_email and '@' in recipient_email:
                        if recipient_email not in email_groups:
                            email_groups[recipient_email] = {
                                'name': info['Dipakai Oleh'],
                                'files': []
                            }
                        email_groups[recipient_email]['files'].append({
                            'excel': output_path,
                            'asset_count': len(assets)
                        })
                
                success += 1
            else:
                print(f"ERROR [{idx}/{len(grouped)}] Failed: {filename}")
        
        # Send emails (grouped by email address)
        if self.send_email and email_groups:
            print(f"\n{'='*60}")
            print("SENDING EMAILS...")
            print(f"{'='*60}\n")
            
            for recipient_email, email_data in email_groups.items():
                person_name = email_data['name']
                files_data = email_data['files']
                
                print(f"Sending to {recipient_email} ({person_name})...")
                print(f"   -> {len(files_data)} file(s) to attach")
                
                if self.send_email_with_attachments(recipient_email, person_name, files_data):
                    email_sent += 1
        
        print(f"\n{'='*60}")
        print(f"OK Completed! {success}/{len(grouped)} files generated")
        print(f"Excel files saved in: {os.path.abspath(self.output_folder)}")
        if self.send_email:
            print(f"Emails sent: {email_sent} (to {len(email_groups)} unique addresses)")
        print(f"{'='*60}")

def auto_detect_csv():
    """Auto-detect CSV file"""
    csv_files = glob.glob("*.csv")
    if csv_files:
        latest = max(csv_files, key=os.path.getmtime)
        print(f"Auto-detected: {latest}")
        return latest
    return None


if __name__ == "__main__":
    # ============================================================
    # EMAIL CONFIG - HARDCODED (Edit di sini untuk setting tetap)
    # ============================================================
    HARDCODED_EMAIL_CONFIG = {
        'smtp_server': 'smtp.gmail.com',      # Ganti dengan SMTP server kamu
        'smtp_port': 587,                      # Port SMTP (587 untuk TLS)
        'sender_email': 'kerjabareng.riska17@gmail.com',  # Ganti dengan email kamu
        'sender_password': 'roig rtuh mwfi nquf'   # Ganti dengan App Password
    }
    
    # Set USE_HARDCODED = True untuk pakai config di atas
    # Set USE_HARDCODED = False untuk input manual setiap kali run
    USE_HARDCODED_EMAIL = True  # <-- Ubah jadi True kalau mau pakai hardcoded
    # ============================================================
    
    print("="*60)
    print("  FORM IT ASSET - EXCEL GENERATOR")
    print("="*60)
    
    TEMPLATE_FILE = 'template_inventaris.xlsx'
    OUTPUT_FOLDER = 'generated_excel'
    
    print("\nSelect Mode:")
    print("  1. CONSOLIDATED - 1 file per person (all assets merged)")
    
    mode = input("\nEnter mode (1) [default=1]: ").strip() or "1"
    
    CSV_FILE = auto_detect_csv()
    
    if CSV_FILE is None:
        print("\nERROR: No CSV file found!")
        print("Place CSV file in the same folder and run again.")
        input("\nPress Enter to exit...")
        exit()
    
    if not os.path.exists(TEMPLATE_FILE):
        print(f"\nERROR: Template not found: {TEMPLATE_FILE}")
        input("\nPress Enter to exit...")
        exit()
    
    # Ask for features
    print("\n" + "="*60)
    print("INSERT IMAGES")
    print("="*60)
    print("  Y - Yes, download and insert images (slower)")
    print("  N - No, skip images (faster)")
    insert_img = input("\nInsert images? (Y/N) [default=Y]: ").strip().upper() or "Y"
    
    print("\n" + "="*60)
    print("SEND EMAIL")
    print("="*60)
    
    email_config = {}  # Initialize here
    
    # Check if hardcoded config is valid
    if USE_HARDCODED_EMAIL and \
       HARDCODED_EMAIL_CONFIG['sender_email'] != 'your-email@gmail.com' and \
       HARDCODED_EMAIL_CONFIG['sender_password'] != 'your-app-password':
        print("✓ Email config found (hardcoded)")
        print(f"  Sender: {HARDCODED_EMAIL_CONFIG['sender_email']}")
        print(f"  Server: {HARDCODED_EMAIL_CONFIG['smtp_server']}:{HARDCODED_EMAIL_CONFIG['smtp_port']}")
        print("\nOptions:")
        print("  Y - Yes, send files to email addresses from CSV")
        print("  N - No, only save locally")
        send_email_input = input("\nSend email? (Y/N) [default=Y]: ").strip().upper() or "Y"
        
        if send_email_input == "Y":
            email_config = HARDCODED_EMAIL_CONFIG.copy()  # Copy config here
    else:
        print("Email config not set or using manual input mode")
        print("\nOptions:")
        print("  Y - Yes, send files (will ask for email config)")
        print("  N - No, only save locally")
        send_email_input = input("\nSend email? (Y/N) [default=N]: ").strip().upper() or "N"
    
    # Only ask for manual input if chose Y but config not set yet
    if send_email_input == "Y" and not email_config:
            
            # Manual input
            print("\n" + "="*60)
            print("EMAIL CONFIGURATION")
            print("="*60)
            print("\nContoh untuk Gmail:")
            print("  SMTP Server: smtp.gmail.com")
            print("  SMTP Port: 587")
            print("  Email: your-email@gmail.com")
            print("  Password: App Password (bukan password Gmail biasa)")
            print("\nCara membuat App Password Gmail:")
            print("  1. Buka Google Account > Security")
            print("  2. Enable 2-Step Verification")
            print("  3. Pilih 'App passwords'")
            print("  4. Generate password untuk 'Mail'")
            print()
            
            email_config['smtp_server'] = input("SMTP Server: ").strip()
            email_config['smtp_port'] = int(input("SMTP Port [587]: ").strip() or "587")
            email_config['sender_email'] = input("Sender Email: ").strip()
            email_config['sender_password'] = input("Email Password/App Password: ").strip()
            
            if not all(email_config.values()):
                print("\nERROR: Email config incomplete! Email sending will be disabled.")
                send_email_input = "N"
    
    # Create generator instance
    generator = SimpleExcelGenerator(
        CSV_FILE, 
        TEMPLATE_FILE, 
        OUTPUT_FOLDER,
        insert_images=(insert_img == "Y"),
        send_email=(send_email_input == "Y"),
        email_config=email_config if send_email_input == "Y" else None
    )
    
    # Run generator
    print("\nRunning CONSOLIDATED mode...\n")
    generator.generate_excel_consolidated()
    
    print("\nDone! Press Enter to exit...")
    input()